from openpyxl import Workbook

wb= Workbook()
ws= wb.active

ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트","총점","성적"))


# 문제조건

# 성적비중 출석 10 퀴즈 10 퀴즈2 10 중간고사 20 기말고사 30 프로젝트 20 총합 100




score = [(1,10,8,5,14,26,12),
 (2,7,3,7,15,24,18),
 (3,9,5,8,8,12,4),
 (4,7,8,7,17,21,18),
 (5,7,8,7,16,25,15),
 (6,3,5,8,8,17,0),
 (7,4,9,10,16,27,18),
 (8,6,6,6,15,19,17),
 (9,10,10,9,19,30,19),
 (10,9,8,8,20,25,20)
 ]



for i in range(0,9) :
    ws.append(score[i])

for i in range(1,9) :
    total_score_row_range = ws[2:i]
     ws[2,i] =  "=SUM(A2:G2)"


# ws["H2"] = "=SUM(A2:G2)" # 30

# for row in ws.values:
#      for cell in row:
#          print(cell)

# for idx,change_quiz_2 in enumerate(ws["D"]) :
#     if idx == 0:
#         continue
#     change_quiz_2.value = 10
# for idx,change_quiz_2 in enumerate(ws["D"]) :
#     if idx == 0:
#         continue
#     change_quiz_2.value = 10

wb.save("quiz.xlsx")

# 퀴즈 문제
# [현재까지 작성된 최종 성적 데이터]
# 학번, 출석, 퀴즈1, 퀴즈2, 중간고사, 기말고사, 프로젝트
# [1,10,8,5,14,26,12
# 2,7,3,7,15,24,18
# 3,9,5,8,8,12,4
# 4,7,8,7,17,21,18
# 5,7,8,7,16,25,15
# 6,3,5,8,8,17,0
# 7,4,9,10,16,27,18
# 8,6,6,6,15,19,17
# 9,10,10,9,19,30,19
# 10,9,8,8,20,25,20
# ]
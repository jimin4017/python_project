from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")

ws = wb.active

A1 = ws["A1"]
B1 = ws["B1"]
C1 = ws["C1"]


#A열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5


# 1행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50

wb.save("sample_heightggg_.xlsx")
from openpyxl import load_workbook
wb = load_workbook("sample_formula.xlsx",data_only=True)
ws = wb.active



for idx,change_quiz_2 in enumerate(ws["D"]) :
    if idx == 0:
        conti
    change_quiz_2.value = 10

     

# wb = load_workbook("sample_formula.xlsx", data_only=True)
# ws = wb.active

# 수식이 아닌 실제 데이터를 가지고 옴
# # evaluate 되지 않은 상태의 데이터는 None 이라고 표시
# for row in ws.values:
#     for cell in row:
#         print(cell)
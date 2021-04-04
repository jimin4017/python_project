from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws.merge_cells("B2:D2") # 셀병합 
ws["B2"].value = " Merged Cell"

wb.save("sample_merge.xlsx")
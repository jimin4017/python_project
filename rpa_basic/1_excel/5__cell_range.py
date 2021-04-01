from openpyxl import Workbook
from random import *

wb= Workbook()

ws = wb.active

ws.append(["번호","영어","수학"])

for i in range(1,21): ## 21개의 데이터 넣기
    ws.append([i,randint(0,100),randint(0,100)])
 
col_B = ws["B"]  ## 영어에서 b칼럼만 가지고 오기
for cell in col_B :
    print(cell.value)

col_range = ws["B:C"] ## 영어 수학 칼럼 함께 가져오기

row_title = ws[1]

for cell in row_title:
    print(cell.value)

from openpyxl.utils.cell import coordinate_from_string
 
row_range = ws[2:6] # 2번재 쭐에서 6번째 줄까지 가지고 오기

for rows in row_range :
    for cell in row_title:
        print(cell.value,end="")
    print()

row_range = ws[2:ws.max_row]

for rows in row_range :
    for cell in rows:
        # print(cell.value,end=" ")    ## end "띄어쓰기"
        # print(cell.coordinate,end=" ") ##  from openpyxl.utils.cell import coordinate_from_string
        xy = coordinate_from_string(cell.coordinate)
        print(xy,end=" ")   ## 위랑 같지만 쉽게 확인 할수 없
        # print(xy[0],end=" ")   ## A열  
        # print(xy[1],end=" ")   ## B열 로 구분 가능 가나다라ㅏ
 

    print()


print(tuple(ws.rows))  ## 한줄씩



print(tuple(ws.columns))  ## 한열씩

# for column in tuple (ws.columns) :
#     print(column[0].value)

for  row in ws.iter_rows() :  #ex) (<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>)
    print(row)

for  row in ws.iter_rows() : #ex) 영어 78
    print(row[1].value)

for row in ws.iter_rows(min_row=1,max_row=5) :  ## 1,5 
    print(row[2].value)                         ## B열




wb.save("sampl_random_append1.xlsx")

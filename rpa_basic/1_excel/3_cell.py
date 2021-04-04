from openpyxl import Workbook
wb = Workbook() # 새 워크북 생성 

ws = wb.active
ws.title = "3_cell_Sheet"


# A1 셀에 1이라는 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3  #쭉쭉들어감

print(ws["A1"]) ## A1에 정보를 출력


print(ws.cell(row=1,column=1).value)

cell_postion = 10

for i in range(1,11) :
    print(ws.cell(row=i,column=i).value)

wb.save("3_cell_Sheet")  ## 파일이름  
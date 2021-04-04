from openpyxl import Workbook

wb = Workbook() # 새 워크북 생성  
ws = wb.active # 현재 활성화된  sheet 가져옴
ws.title = "jiminSheet_test_A1"  ## 이름 변경

ws.sheet_properties.tabColor = "ff66ff" ## rgb 값 가져와서 표시하는거

# sheet, mysheet, your sheet index값 위치 설정
ws1 = wb.create_sheet("your sheet") # 주어진  이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet",2) # ("이름,2") 2번째에 2번이름 시트로 생성

new_ws = wb["NewSheet"]

new_ws["A1"] = "안녕" 

target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"


wb.save("sample2_change title.xlsx")  ## 파일이름


wb.close()